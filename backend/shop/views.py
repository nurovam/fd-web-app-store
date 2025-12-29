from __future__ import annotations

from typing import Any, Dict, List
from django.contrib.auth import get_user_model
from django.db import transaction
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer

from openpyxl import load_workbook

from .models import Category, Product, Cart, CartItem, Order, ImportJob
from .serializers import (
    CategorySerializer,
    ProductSerializer,
    CartSerializer,
    CartItemSerializer,
    OrderSerializer,
    UserSerializer,
    ImportJobSerializer,
)

User = get_user_model()


class RegistrationView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = UserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)


class LoginView(TokenObtainPairView):
    serializer_class = TokenObtainPairSerializer
    permission_classes = [permissions.AllowAny]


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    lookup_field = "slug"


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.select_related("category").all()
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    filterset_fields = ["category__slug", "is_featured"]
    search_fields = ["title", "sku", "description"]
    ordering_fields = ["price", "created_at"]


class CartViewSet(viewsets.ViewSet):
    permission_classes = [permissions.AllowAny]

    def _get_cart(self, request) -> Cart:
        if request.user and request.user.is_authenticated:
            cart, _ = Cart.objects.get_or_create(user=request.user)
        else:
            session_key = request.data.get("session_key") or request.query_params.get("session_key", "")
            cart, _ = Cart.objects.get_or_create(session_key=session_key)
        return cart

    def list(self, request):
        cart = self._get_cart(request)
        serializer = CartSerializer(cart)
        return Response(serializer.data)

    @action(detail=False, methods=["post"])
    def add(self, request):
        cart = self._get_cart(request)
        serializer = CartItemSerializer(data=request.data, context={"cart": cart})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        cart.refresh_from_db()
        return Response(CartSerializer(cart).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["patch"])
    def update_quantity(self, request, pk=None):
        cart = self._get_cart(request)
        item = get_object_or_404(CartItem, pk=pk, cart=cart)
        serializer = CartItemSerializer(instance=item, data=request.data, partial=True, context={"cart": cart})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        cart.refresh_from_db()
        return Response(CartSerializer(cart).data)

    @action(detail=True, methods=["delete"])
    def remove(self, request, pk=None):
        cart = self._get_cart(request)
        item = get_object_or_404(CartItem, pk=pk, cart=cart)
        item.delete()
        cart.refresh_from_db()
        return Response(CartSerializer(cart).data)


class OrderViewSet(viewsets.ModelViewSet):
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Order.objects.filter(user=self.request.user).prefetch_related("items__product")

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ImportProductsView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "Файл не найден"}, status=status.HTTP_400_BAD_REQUEST)

        job = ImportJob.objects.create(
            created_by=request.user if request.user.is_authenticated else None,
            original_filename=uploaded_file.name,
        )

        try:
            report = self._process_workbook(uploaded_file)
            job.status = ImportJob.Status.COMPLETED
            job.processed_rows = report["created"] + report["updated"]
            job.report = report
            job.save(update_fields=["status", "processed_rows", "report"])
            return Response(ImportJobSerializer(job).data)
        except Exception as exc:  # pragma: no cover - safety net for user uploads
            job.status = ImportJob.Status.FAILED
            job.error_message = str(exc)
            job.save(update_fields=["status", "error_message"])
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    def _process_workbook(self, uploaded_file) -> Dict[str, Any]:
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        required_headers = {"name", "sku", "price", "inventory"}
        if not required_headers.issubset(set(headers)):
            missing = ", ".join(required_headers - set(headers))
            raise ValueError(f"Отсутствуют обязательные столбцы: {missing}")

        header_index = {header: idx for idx, header in enumerate(headers)}
        created = updated = 0
        errors: List[str] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            payload = {
                "title": row[header_index.get("name")],
                "sku": row[header_index.get("sku")],
                "price": row[header_index.get("price")],
                "inventory": row[header_index.get("inventory")] or 0,
                "description": row[header_index.get("description")] if header_index.get("description") is not None else "",
                "hero_image": row[header_index.get("image_url")] if header_index.get("image_url") is not None else "",
                "category_name": row[header_index.get("category")] if header_index.get("category") is not None else "",
            }

            if not payload["title"] or not payload["sku"]:
                errors.append(f"Пропущены данные в строке: {row}")
                continue

            with transaction.atomic():
                category = None
                if payload["category_name"]:
                    category, _ = Category.objects.get_or_create(name=payload["category_name"])
                product, created_flag = Product.objects.update_or_create(
                    sku=payload["sku"],
                    defaults={
                        "title": payload["title"],
                        "price": payload["price"],
                        "inventory": payload["inventory"],
                        "description": payload["description"],
                        "hero_image": payload["hero_image"] or "",
                        "category": category,
                    },
                )
                if created_flag:
                    created += 1
                else:
                    updated += 1

        return {"created": created, "updated": updated, "errors": errors}
